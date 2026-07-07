"""Import active grants and discretionary accounts from a grant-report workbook.

Works with any .xlsx report that has a sheet containing a header row with at
least these column names (case-insensitive, any order, other columns
ignored): "Project Name", "PTA", "Award End Date", "Budget", "Balance". That
sheet is located automatically by scanning every sheet's first few rows for
a matching header - no fixed sheet name or file name is assumed. A single
non-empty cell with the rest of the row blank (e.g. "Research Awards",
"Discretionary Accounts") is treated as a section label for the rows below
it; any section whose label contains "discretionary" is exempt from the
active/expired end-date filter, since discretionary and gift accounts don't
really "expire" the way an award does.

Optionally, if some sheet has a header row with "Project Number" and an
expenditure-date column (e.g. "Expenditure Item Date"), the earliest
transaction date per project number (the PTA value's first dot-segment) is
used to infer each grant's start date. Grants with no match fall back to an
estimated start date (end date minus --estimate-years years), flagged in
their notes.

The report format has no sponsor column, so `sponsor` is left blank on
import - fill it in via the app afterwards if useful.

For each imported grant, one lump-sum transaction is recorded (amount =
budget - balance, dated --as-of-date or today) so the balance shown in the
app matches the sheet's reported Balance column exactly.

Safe to re-run: each imported grant is tagged with its PTA number in its
notes, and rows already tagged with that PTA are skipped.

Run:
  ./venv/bin/python import_from_excel.py REPORT.xlsx --faculty "Jane Smith"
  ./venv/bin/python import_from_excel.py REPORT.xlsx --faculty "Jane Smith" --dry-run
  ./venv/bin/python import_from_excel.py REPORT.xlsx --faculty "Jane Smith" --as-of-date 2026-05-31
"""
import argparse
import datetime
import re
import sqlite3

import openpyxl

from app import DATA_DIR, SCHEMA_PATH, slugify

REPORT_REQUIRED_COLUMNS = {"project name", "pta", "award end date", "budget", "balance"}
LEDGER_REQUIRED_COLUMNS = {"project number", "expenditure item date"}
HEADER_SEARCH_ROWS = 50  # header rows are always near the top; cap the scan on huge data sheets


def normalize_header(row):
    """{lowercased column label: index} for the non-empty cells in a row."""
    return {str(c).strip().lower(): i for i, c in enumerate(row) if c is not None}


def find_header(wb, required_columns):
    """Scan every sheet's first HEADER_SEARCH_ROWS rows for a row whose labels
    are a superset of required_columns. Several sheets can coincidentally share
    a column layout (e.g. a small reconciliation tab vs. the real ledger), so
    among matches, prefer the sheet with the most rows. Returns
    (sheet_name, row_index, col_index) or (None, None, None) if nothing matches."""
    candidates = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_index, row in enumerate(ws.iter_rows(max_row=HEADER_SEARCH_ROWS, values_only=True), start=1):
            labels = normalize_header(row)
            if required_columns.issubset(labels.keys()):
                candidates.append((ws.max_row, sheet_name, row_index, labels))
                break
    if not candidates:
        return None, None, None
    candidates.sort(key=lambda c: c[0], reverse=True)
    _, sheet_name, row_index, labels = candidates[0]
    return sheet_name, row_index, labels


def project_number(pta):
    return str(pta).split(".")[0].strip()


def build_start_date_lookup(wb):
    """project number (PTA's first dot-segment) -> earliest expenditure date found
    in a ledger sheet, if one exists in this workbook. {} if no such sheet is found."""
    sheet_name, header_row, col = find_header(wb, LEDGER_REQUIRED_COLUMNS)
    if sheet_name is None:
        return {}
    ws = wb[sheet_name]
    proj_col = col["project number"]
    date_col = col["expenditure item date"]
    lookup = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if proj_col >= len(row) or date_col >= len(row):
            continue
        proj, date = row[proj_col], row[date_col]
        if proj is None or not isinstance(date, datetime.datetime):
            continue
        proj = str(proj).strip()
        d = date.date()
        if proj not in lookup or d < lookup[proj]:
            lookup[proj] = d
    return lookup


def parse_report_rows(wb, sheet_name, col):
    """Yield (section, name, pta, end_date, budget, balance, notes) for each data row.

    Scans the whole sheet from the top (not just below the first header row we
    detected) since a section label - e.g. "Research Awards" - can appear
    before the header row it introduces, and the header row itself can repeat
    once per section."""
    ws = wb[sheet_name]
    section = None
    notes_col = col.get("notes")
    for row in ws.iter_rows(values_only=True):
        labels = normalize_header(row)
        if REPORT_REQUIRED_COLUMNS.issubset(labels.keys()):
            continue  # a header row (this report format repeats it per section)

        name = row[col["project name"]] if col["project name"] < len(row) else None
        if name is None or not str(name).strip():
            continue
        name = re.sub(r"\s+", " ", str(name)).strip()

        rest_of_row = [v for i, v in enumerate(row) if i != col["project name"]]
        if all(v is None for v in rest_of_row):
            section = name  # a section-label row: only the first column is populated
            continue

        pta = row[col["pta"]] if col["pta"] < len(row) else None
        if not pta:
            continue  # category label row, no actual account
        end_date = row[col["award end date"]] if col["award end date"] < len(row) else None
        budget = row[col["budget"]] if col["budget"] < len(row) else None
        balance = row[col["balance"]] if col["balance"] < len(row) else None
        notes = row[notes_col] if notes_col is not None and notes_col < len(row) else None

        budget = budget or 0
        balance = balance or 0
        if budget == 0 and balance == 0:
            continue  # dead/zeroed-out row

        yield section, name, str(pta).strip(), end_date, budget, balance, notes


def build_grant_records(wb, sheet_name, col, today, include_expired, estimate_years):
    start_lookup = build_start_date_lookup(wb)
    records = []
    for section, name, pta, end_date, budget, balance, notes in parse_report_rows(wb, sheet_name, col):
        if not isinstance(end_date, datetime.datetime):
            continue
        end_date = end_date.date()
        is_discretionary = section is not None and "discretionary" in section.lower()
        if not is_discretionary and not include_expired and end_date < today:
            continue  # not active

        proj = project_number(pta)
        start_date = start_lookup.get(proj)
        estimated_start = False
        if start_date is None:
            start_date = end_date - datetime.timedelta(days=365 * estimate_years)
            estimated_start = True

        note_parts = [f"PTA {pta}."]
        if notes:
            note_parts.append(str(notes).strip())
        if estimated_start:
            note_parts.append(
                f"[start date estimated: no ledger history found, set to {estimate_years} years before end date]"
            )
        full_notes = " ".join(note_parts)

        records.append({
            "section": section,
            "name": name,
            "pta": pta,
            "total_amount_cents": round(budget * 100),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "notes": full_notes,
            "spent_cents": round((budget - balance) * 100),
        })
    return records


def existing_ptas(db):
    ptas = set()
    for (notes,) in db.execute("SELECT notes FROM grants WHERE notes LIKE 'PTA %'"):
        m = re.match(r"PTA (\S+)\.", notes or "")
        if m:
            ptas.add(m.group(1))
    return ptas


def open_db(faculty_name):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    slug = slugify(faculty_name)
    path = DATA_DIR / f"{slug}.db"
    db = sqlite3.connect(path)
    db.execute("PRAGMA foreign_keys = ON")
    if not path.exists() or path.stat().st_size == 0:
        db.executescript(SCHEMA_PATH.read_text())
    else:
        db.execute("SELECT 1 FROM grants LIMIT 1")  # sanity check schema exists
    return db, path


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("xlsx", help="Path to the report workbook (.xlsx)")
    parser.add_argument("--faculty", required=True, help='Faculty database to import into, e.g. "Jane Smith"')
    parser.add_argument("--as-of-date", help="YYYY-MM-DD date to record imported transactions as of (default: today)")
    parser.add_argument("--estimate-years", type=int, default=3,
                         help="Years before end date to use as an estimated start date when no ledger history is found (default: 3)")
    parser.add_argument("--include-expired", action="store_true",
                         help="Also import awards whose end date has passed (discretionary accounts are always included)")
    parser.add_argument("--dry-run", action="store_true", help="Print what would be imported, write nothing")
    args = parser.parse_args()

    as_of_date = datetime.date.fromisoformat(args.as_of_date) if args.as_of_date else datetime.date.today()
    today = datetime.date.today()

    print(f"Loading {args.xlsx} ...")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True)

    sheet_name, header_row, col = find_header(wb, REPORT_REQUIRED_COLUMNS)
    if sheet_name is None:
        raise SystemExit(
            "Couldn't find a sheet with a header row containing "
            f"{sorted(REPORT_REQUIRED_COLUMNS)}. This script expects a report in that format."
        )
    print(f"Using report sheet {sheet_name!r} (header on row {header_row}).")

    records = build_grant_records(wb, sheet_name, col, today, args.include_expired, args.estimate_years)
    print(f"Found {len(records)} active/discretionary accounts with financial substance.")

    db, path = open_db(args.faculty)
    already = existing_ptas(db)

    inserted, skipped = 0, 0
    for r in records:
        if r["pta"] in already:
            skipped += 1
            continue
        print(f"  [{r['section']}] {r['name']!r} ({r['pta']}) "
              f"budget=${r['total_amount_cents']/100:,.2f} "
              f"{r['start_date']}..{r['end_date']}")
        if not args.dry_run:
            cur = db.execute(
                """INSERT INTO grants (name, sponsor, total_amount_cents, start_date, end_date, overhead_rate_bps, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (r["name"], "", r["total_amount_cents"], r["start_date"], r["end_date"], 0, r["notes"]),
            )
            grant_id = cur.lastrowid
            if r["spent_cents"]:
                db.execute(
                    "INSERT INTO transactions (grant_id, date, amount_cents, description) VALUES (?, ?, ?, ?)",
                    (grant_id, as_of_date.isoformat(), r["spent_cents"],
                     f"Expenditures + encumbrances as of {as_of_date.isoformat()} (imported from {args.xlsx})"),
                )
        inserted += 1

    if args.dry_run:
        print(f"\nDry run: would insert {inserted}, skip {skipped} (already imported). No changes written.")
    else:
        db.commit()
        print(f"\nInserted {inserted}, skipped {skipped} (already imported). Wrote to {path}")
    db.close()


if __name__ == "__main__":
    main()
